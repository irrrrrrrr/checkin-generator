import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import openpyxl
import os
import time

temp_path = '../temp/template.xlsx'


def process_excel_file_pandas_openpyxl(desktop_path):
    try:
        file_path = os.path.join(desktop_path, "off_time.xlsx")
        print(f"Loading workbook from {temp_path}")

        # 使用 pandas 读取 Excel 文件
        df = pd.read_excel(temp_path, header=None)

        # 处理换行符
        df = df.applymap(lambda x: x.split('\n', 1)[1] if isinstance(x, str) and '\n' in x else x)

        # 使用 openpyxl 读取原始 Excel 文件
        wb = openpyxl.load_workbook(temp_path)
        sheet = wb.active

        # 清空原始工作表的数据
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.value = None

        # 将 pandas DataFrame 写回到 openpyxl 工作表，保持格式
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        # 保存处理后的 Excel 文件
        wb.save(temp_path)
        print(f"Workbook saved to {temp_path}")

        if os.path.exists(temp_path):
            shutil.copy(temp_path, file_path)
            print(f"Copied {temp_path} to {file_path}")
        else:
            print(f"Temp file does not exist: {temp_path}")
    except Exception as e:
        print(f"Error in process_excel_file: {e}")
        import traceback
        traceback.print_exc()


def process_excel_file_openpyxl(desktop_path):
    try:
        file_path = os.path.join(desktop_path, "off_time.xlsx")
        print(f"Loading workbook from {temp_path}")
        wb = openpyxl.load_workbook(temp_path)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.value and '\n' in cell.value:
                    a, b = cell.value.split('\n', 1)
                    cell.value = b

        wb.save(temp_path)
        print(f"Workbook saved to {temp_path}")

        if os.path.exists(temp_path):
            shutil.copy(temp_path, file_path)
            print(f"Copied {temp_path} to {file_path}")
        else:
            print(f"Temp file does not exist: {temp_path}")
    except Exception as e:
        print(f"Error in process_excel_file: {e}")
        import traceback
        traceback.print_exc()


desktop_path = "C:\\Users\\guess\\Desktop"

start_time = time.time()
process_excel_file_openpyxl(desktop_path)
print(f"openpyxl method took {time.time() - start_time} seconds")

start_time = time.time()
process_excel_file_pandas_openpyxl(desktop_path)
print(f"pandas + openpyxl method took {time.time() - start_time} seconds")