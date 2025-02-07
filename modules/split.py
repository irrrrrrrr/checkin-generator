import shutil
import openpyxl
import subprocess
import os


temp_path = './temp/template.xlsx'


def process_excel_file(desktop_path):
    try:
        file_path = f"{desktop_path}\\off_time.xlsx"
        print(f"Loading workbook from {temp_path}")
        wb = openpyxl.load_workbook(temp_path)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.value and '\n' in cell.value:
                    a, b = cell.value.split('\n', 1)
                    cell.value = b

        wb.save(temp_path)
        print(f"Workbook saved to {temp_path}")

        if os.path.exists(temp_path):
            shutil.copy(temp_path, file_path)
            print(f"Copied {temp_path} to {file_path}")
            run(file_path)
        else:
            print(f"Temp file does not exist: {temp_path}")
    except Exception as e:
        print(f"Error in process_excel_file: {e}")
        import traceback
        traceback.print_exc()


def run(file_dir):
    try:
        print(f"Running Excel with {file_dir}")
        excel_path = "C:\\Program Files\\Microsoft Office\\root\\Office16\\EXCEL.EXE"
        subprocess.run([excel_path, file_dir])
    except Exception as e:
        print(f"Error running Excel: {e}")
        import traceback
        traceback.print_exc()


def copy():
    try:
        print(f"Copying template.xlsx to {temp_path}")
        shutil.copy('template.xlsx', temp_path)
        print("Copy complete")
    except Exception as e:
        print(f"Error copying template: {e}")
        import traceback
        traceback.print_exc()


def remove_temp():
    try:
        print(f"Deleting temp files")
        print("Delete complete")
        os.remove(temp_path)
    except:
        pass
