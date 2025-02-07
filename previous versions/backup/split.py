import shutil  
import openpyxl  

shutil.copy('C:\\Users\\guess\\Desktop\\previous versions\\backup\\template.xlsx', 'C:\\Users\\guess\\Desktop\\off_time.xlsx')
wb = openpyxl.load_workbook('C:\\Users\\guess\\Desktop\\off_time.xlsx')  
sheet = wb.active  

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):  
    for cell in row:  
        if cell.value and '\n' in cell.value:  
            a, b = cell.value.split('\n', 1)  
            cell.value = b  

wb.save('C:\\Users\\guess\\Desktop\\off_time.xlsx')